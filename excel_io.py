"""
Excel helpers: pandas reads + openpyxl writes to the SAME file.

We use pandas to read tabular data (as requested) and openpyxl to update only
the "URL Status" cells so other sheets/formatting are preserved as much as possible.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

LINKEDIN_URL_HEADER = "LinkedIn URL"
URL_STATUS_HEADER = "URL Status"


@dataclass(frozen=True)
class ExcelContext:
    path: Path
    sheet_name: str
    linkedin_col: int  # 1-based column index
    status_col: int  # 1-based column index


def _normalize_header(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _find_header_column(ws: Worksheet, header_name: str, header_row: int = 1) -> int | None:
    for col in range(1, ws.max_column + 1):
        if _normalize_header(ws.cell(row=header_row, column=col).value) == header_name:
            return col
    return None


def validate_linkedin_column_with_pandas(path: Path, sheet_name: str | int = 0) -> pd.DataFrame:
    """
    Load the sheet with pandas and ensure a 'LinkedIn URL' column exists (header match after strip).
    """
    df = pd.read_excel(path, engine="openpyxl", sheet_name=sheet_name, header=0)
    df.columns = [_normalize_header(c) for c in df.columns]
    if LINKEDIN_URL_HEADER not in df.columns:
        raise ValueError(
            f'Missing "{LINKEDIN_URL_HEADER}" column in the first sheet. '
            f'Found columns: {list(df.columns)[:25]}{"..." if len(df.columns) > 25 else ""}'
        )
    return df


def prepare_workbook_columns(path: Path, sheet_index: int = 0) -> ExcelContext:
    """
    Open the workbook with openpyxl, locate/create URL Status column, return column indices.
    """
    wb = load_workbook(path, read_only=False, data_only=False)
    ws = wb.worksheets[sheet_index]
    sheet_name = ws.title

    linkedin_col = _find_header_column(ws, LINKEDIN_URL_HEADER, header_row=1)
    if linkedin_col is None:
        wb.close()
        raise ValueError(f'Row 1 must contain a header cell exactly named "{LINKEDIN_URL_HEADER}".')

    status_col = _find_header_column(ws, URL_STATUS_HEADER, header_row=1)
    if status_col is None:
        status_col = ws.max_column + 1
        ws.cell(row=1, column=status_col, value=URL_STATUS_HEADER)

    wb.save(path)
    wb.close()

    return ExcelContext(path=path, sheet_name=sheet_name, linkedin_col=linkedin_col, status_col=status_col)


def read_url_for_excel_row(
    *,
    df: pd.DataFrame,
    excel_row: int,
    ctx: ExcelContext,
    wb: Workbook,
) -> str:
    """
    Return the raw cell string for LinkedIn URL on this 1-based Excel row.

    Row 1 is read from the worksheet (header row). Rows >= 2 use pandas iloc mapping
    when the sheet's first row is the table header (pandas header=0).
    """
    ws = wb[ctx.sheet_name]
    if excel_row < 1:
        raise ValueError("excel_row must be >= 1")

    if excel_row == 1:
        raw = ws.cell(row=1, column=ctx.linkedin_col).value
        return "" if raw is None else str(raw)

    # pandas row index when Excel row 1 is the header row
    pandas_idx = excel_row - 2
    if pandas_idx < 0 or pandas_idx >= len(df):
        raw = ws.cell(row=excel_row, column=ctx.linkedin_col).value
        return "" if raw is None else str(raw)

    series_val = df.iloc[pandas_idx][LINKEDIN_URL_HEADER]
    if pd.isna(series_val):
        return ""
    return str(series_val)


def write_status_cell(*, wb: Workbook, ctx: ExcelContext, excel_row: int, status: str) -> None:
    ws = wb[ctx.sheet_name]
    ws.cell(row=excel_row, column=ctx.status_col, value=status)


def save_workbook(wb: Workbook, path: Path) -> None:
    wb.save(path)
